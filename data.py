import openpyxl as xl

def get_aircraft_data():
    data_wb = xl.load_workbook(filename="aircraft-data.xlsx", data_only=True)
    ws_names = data_wb.sheetnames
    data_list = []
    vsp_data = {}
    for active_ws in ws_names:
        ws = data_wb[active_ws]
        data_dict = {}
        if active_ws != "VSP Dados":
            for i in range(1, ws.max_row + 1):
                data_dict[ws.cell(row=i, column=1).value] = ws.cell(row=i, column=2).value
        else:
            for i in range(1, ws.max_column + 1):
                for j in range(1, ws.max_row + 1):
                    data_dict[
                        f"{ws.cell(row=i, column=1).value}_{ws.cell(row=1, column=j).value}"
                    ] = ws.cell(row=i, column=j).value

        data_list.append(data_dict)
    
    return data_list
